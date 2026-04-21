"""Google Maps Routes API client for traffic monitoring.

Two entry points:
    * ``gmaps_typical`` — one-shot fetch of Google's *typical* traffic for one
      future day at a fixed cadence. Uses the fact that Routes API with a
      future ``departureTime`` returns Google's historical-pattern-based
      prediction. Gives us a reference curve without waiting for live polling
      to accumulate.
    * ``gmaps_poll`` — live daemon that calls the same endpoint with
      ``departureTime=now`` and appends to a rolling ndjson log.

The API key is read from ``GOOGLE_MAPS_API_KEY`` (either the shell environment
or a ``.env`` file at the repo root). Never pass the key on the command line
or hard-code it.

Handbook note: §11 mandates open-source / no paid lock-in. Google Maps is a
pragmatic exception, documented in ``phase1-sandbox/methodology.md``, because
no free API exposes real Amman traffic at the granularity Phase 1 needs to
validate against.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import httpx
import yaml


ROUTES_URL = "https://routes.googleapis.com/directions/v2:computeRoutes"
# Field mask restricts the response to what we need — required by Routes API v2.
# routes.description gives the human-readable "Via X Street / Y Street" summary
# that we surface as the corridor's street name.
FIELD_MASK = ("routes.duration,routes.staticDuration,routes.distanceMeters,"
              "routes.description")

# Jordan time — fixed UTC+3 since 2022 (Jordan dropped DST)
AMMAN_TZ = timezone(timedelta(hours=3))


@dataclass
class Corridor:
    id: str
    bearing_deg: float
    origin_lat: float
    origin_lon: float
    dest_lat: float
    dest_lon: float


def _load_env(repo_root: Path) -> None:
    """Minimal .env loader. Only imports KEY=VALUE lines, ignores quotes and
    comments. Skipped silently if the file doesn't exist."""
    env_path = repo_root / ".env"
    if not env_path.exists():
        return
    try:
        for raw in env_path.read_text().splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = val
    except OSError:
        pass


def load_routes(config_path: Path) -> tuple[dict, list[Corridor]]:
    cfg = yaml.safe_load(config_path.read_text())
    corridors = [
        Corridor(
            id=c["id"],
            bearing_deg=float(c.get("bearing_deg", 0)),
            origin_lat=float(c["origin"]["lat"]),
            origin_lon=float(c["origin"]["lon"]),
            dest_lat=float(c["destination"]["lat"]),
            dest_lon=float(c["destination"]["lon"]),
        )
        for c in cfg["corridors"]
    ]
    return cfg, corridors


def call_routes(
    corr: Corridor,
    departure: datetime | None,
    api_key: str,
    client: httpx.Client,
) -> dict:
    """Single POST to Routes API. Returns a flat dict for ndjson writing."""
    body: dict = {
        "origin":      {"location": {"latLng": {"latitude": corr.origin_lat,
                                                "longitude": corr.origin_lon}}},
        "destination": {"location": {"latLng": {"latitude": corr.dest_lat,
                                                "longitude": corr.dest_lon}}},
        "travelMode":        "DRIVE",
        "routingPreference": "TRAFFIC_AWARE",
    }
    if departure is not None:
        body["departureTime"] = departure.astimezone(timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ")

    headers = {
        "Content-Type":         "application/json",
        "X-Goog-Api-Key":       api_key,
        "X-Goog-FieldMask":     FIELD_MASK,
    }
    resp = client.post(ROUTES_URL, json=body, headers=headers, timeout=10.0)
    # Routes API returns an empty routes list (not a 4xx) when the point is
    # unroutable. Flag that explicitly.
    if resp.status_code != 200:
        return {
            "ok": False,
            "error": f"http {resp.status_code}: {resp.text[:200]}",
            "corridor": corr.id,
        }
    data = resp.json()
    routes = data.get("routes") or []
    if not routes:
        return {
            "ok": False,
            "error": "no route found (unroutable point?)",
            "corridor": corr.id,
            "raw": data,
        }
    r0 = routes[0]
    dur_s    = _duration_to_seconds(r0.get("duration",       "0s"))
    static_s = _duration_to_seconds(r0.get("staticDuration", "0s"))
    distance_m = float(r0.get("distanceMeters", 0))
    congestion = (dur_s / static_s) if static_s > 0 else None
    # Normalize the "Via X / Y" text: keep it short and strip the "Via " prefix
    description = (r0.get("description") or "").strip()
    if description.lower().startswith("via "):
        description = description[4:]
    return {
        "ok": True,
        "corridor":          corr.id,
        "bearing_deg":       corr.bearing_deg,
        "street_name":       description or None,
        "duration_s":        dur_s,
        "static_duration_s": static_s,
        "distance_m":        distance_m,
        "congestion_ratio":  round(congestion, 4) if congestion is not None else None,
        "speed_kmh":         (round(distance_m / dur_s * 3.6, 2)
                              if dur_s > 0 else None),
        "static_speed_kmh":  (round(distance_m / static_s * 3.6, 2)
                              if static_s > 0 else None),
    }


def _duration_to_seconds(val: str | int | float) -> float:
    """Routes API returns duration as a string like '328s'. Handle either."""
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.endswith("s"):
        s = s[:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _congestion_label(ratio: float | None) -> str:
    if ratio is None:
        return "unknown"
    if ratio < 1.15:  return "free"
    if ratio < 1.40:  return "light"
    if ratio < 1.80:  return "moderate"
    if ratio < 2.40:  return "heavy"
    return "jam"


# ─── Subcommand: typical ────────────────────────────────────────────────────
def cmd_typical(args: argparse.Namespace) -> int:
    repo_root = Path(__file__).resolve().parents[4]
    _load_env(repo_root)
    api_key = os.environ.get("GOOGLE_MAPS_API_KEY")
    if not api_key:
        print("[gmaps] GOOGLE_MAPS_API_KEY not set (check .env)", file=sys.stderr)
        return 2

    cfg, corridors = load_routes(args.config)
    target_date = date.fromisoformat(args.date)
    interval_min = int(args.interval_min)
    slots = [
        datetime.combine(target_date, datetime.min.time(),
                         tzinfo=AMMAN_TZ) + timedelta(minutes=interval_min * i)
        for i in range((24 * 60) // interval_min)
    ]
    total_calls = len(slots) * len(corridors)
    print(f"[gmaps] typical fetch: date={target_date} "
          f"({target_date.strftime('%A')} Amman-local)  "
          f"corridors={len(corridors)}  slots={len(slots)}  "
          f"total_calls={total_calls}", file=sys.stderr)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    rows: list[dict] = []
    with httpx.Client() as client:
        for slot in slots:
            for corr in corridors:
                res = call_routes(corr, slot, api_key, client)
                row = {
                    "fetched_at":    now_iso,
                    "departure_local": slot.strftime("%Y-%m-%dT%H:%M:%S%z"),
                    "departure_utc":   slot.astimezone(timezone.utc)
                                           .strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "local_hour":    slot.hour + slot.minute / 60,
                    "site_id":       cfg.get("site_id", "SITE-GMAPS"),
                    **res,
                }
                if row.get("ok"):
                    row["congestion_label"] = _congestion_label(row["congestion_ratio"])
                rows.append(row)
                # Light politeness; stay under 200 QPS default quota
                time.sleep(0.05)
            print(f"  slot {slot.strftime('%H:%M')} done "
                  f"({len(rows)}/{total_calls})", file=sys.stderr)

    # Write ndjson (human-legible) AND parquet (for viewer chart)
    with args.out.open("w") as fh:
        for r in rows:
            fh.write(json.dumps(r) + "\n")
    print(f"[gmaps] wrote {len(rows)} rows → {args.out}", file=sys.stderr)

    # Parquet + XLSX siblings for the viewer chart + Excel review
    try:
        import pandas as pd
        import pyarrow.parquet as pq
        import pyarrow as pa
        df = pd.DataFrame.from_records([r for r in rows if r.get("ok")])
        if not df.empty:
            parquet_path = args.out.with_suffix(".parquet")
            pq.write_table(pa.Table.from_pandas(df, preserve_index=False),
                           parquet_path, compression="zstd")
            print(f"[gmaps] wrote {len(df)} rows → {parquet_path}", file=sys.stderr)

            xlsx_path = args.out.with_suffix(".xlsx")
            _write_typical_xlsx(df, xlsx_path)
            print(f"[gmaps] wrote xlsx → {xlsx_path}", file=sys.stderr)
    except Exception as exc:  # noqa: BLE001
        print(f"[gmaps] parquet/xlsx write skipped: {exc}", file=sys.stderr)
    return 0


def _write_typical_xlsx(df: "pd.DataFrame", dst: Path) -> None:
    """Emit a reviewer-friendly Excel with a title row, one pivot per metric
    (rows = Amman local time, columns = <corridor> corridor), plus a raw-rows
    sheet with both local and UTC timestamps. Sheet names avoid Excel's
    disallowed characters (/ ? * [ ] :)."""
    import pandas as pd  # noqa: PLC0415
    from openpyxl import load_workbook  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill  # noqa: PLC0415

    df = df.copy()
    df["Local time (Amman)"] = df["departure_local"].str[11:16]
    df["UTC time"]           = df["departure_utc"].str[11:16]
    target_date = df["departure_local"].str[:10].iloc[0]
    weekday = pd.Timestamp(target_date).day_name()

    def _pivot(metric: str) -> "pd.DataFrame":
        w = df.pivot_table(index="Local time (Amman)",
                           columns="corridor",
                           values=metric,
                           aggfunc="first").sort_index()
        w.columns = [f"{c} corridor" for c in w.columns]
        return w

    sheets = {
        "Congestion ratio":    _pivot("congestion_ratio"),
        "Speed kmh":           _pivot("speed_kmh"),
        "Free-flow speed kmh": _pivot("static_speed_kmh"),
        "Duration s":          _pivot("duration_s"),
        "Distance m":          _pivot("distance_m"),
    }
    raw_cols = ["Local time (Amman)", "UTC time", "corridor", "bearing_deg",
                "ok", "congestion_ratio", "congestion_label",
                "speed_kmh", "static_speed_kmh",
                "duration_s", "static_duration_s", "distance_m",
                "departure_local", "departure_utc"]
    df_raw = df[[c for c in raw_cols if c in df.columns]].sort_values(
        ["Local time (Amman)", "corridor"])

    with pd.ExcelWriter(dst, engine="openpyxl") as xw:
        for name, wide in sheets.items():
            wide.to_excel(xw, sheet_name=name, startrow=2)
        df_raw.to_excel(xw, sheet_name="Raw rows", index=False)

    wb = load_workbook(dst)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="3E444D", end_color="3E444D",
                           fill_type="solid")
    title_font = Font(bold=True, size=13)
    title = f"Typical traffic · {target_date} ({weekday}) · Amman · SITE-GMAPS"
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name != "Raw rows":
            ws.cell(row=1, column=1, value=title).font = title_font
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=6)
            ws.freeze_panes = "B4"
            for col in range(1, 6):
                c = ws.cell(row=3, column=col)
                c.font = hdr_font; c.fill = hdr_fill
            ws.column_dimensions["A"].width = 20
            for letter in ("B", "C", "D", "E"):
                ws.column_dimensions[letter].width = 14
        else:
            ws.freeze_panes = "A2"
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col)
                c.font = hdr_font; c.fill = hdr_fill
            for letter in ("A","B","C","D","E","F","G","H","I","J","K","L","M","N"):
                ws.column_dimensions[letter].width = 18
    wb.save(dst)


# ─── Subcommand: poll ───────────────────────────────────────────────────────
def cmd_poll(args: argparse.Namespace) -> int:
    repo_root = Path(__file__).resolve().parents[4]
    _load_env(repo_root)
    api_key = os.environ.get("GOOGLE_MAPS_API_KEY")
    if not api_key:
        print("[gmaps] GOOGLE_MAPS_API_KEY not set (check .env)", file=sys.stderr)
        return 2

    cfg, corridors = load_routes(args.config)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    interval_s = float(args.interval_s)
    print(f"[gmaps] polling {len(corridors)} corridors every {interval_s:.0f}s "
          f"→ {args.out}", file=sys.stderr)

    with httpx.Client() as client:
        while True:
            now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            with args.out.open("a") as fh:
                for corr in corridors:
                    res = call_routes(corr, None, api_key, client)  # now
                    row = {
                        "timestamp": now_iso,
                        "mode":      "live",
                        "site_id":   cfg.get("site_id", "SITE-GMAPS"),
                        **res,
                    }
                    if row.get("ok"):
                        row["congestion_label"] = _congestion_label(
                            row["congestion_ratio"])
                    fh.write(json.dumps(row) + "\n")
                    time.sleep(0.05)
            if args.once:
                return 0
            time.sleep(interval_s)


# ─── CLI ────────────────────────────────────────────────────────────────────
def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(prog="gmaps", description=__doc__.splitlines()[0])
    sub = p.add_subparsers(dest="subcmd", required=True)

    CONFIG_DEFAULT = Path("phase1-sandbox/configs/gmaps_routes.yml")

    p_typ = sub.add_parser("typical", help="One-shot fetch of Google's typical-traffic prediction for a full day.")
    p_typ.add_argument("--config",       type=Path, default=CONFIG_DEFAULT)
    p_typ.add_argument("--date",         type=str, required=True,
                       help="Target date YYYY-MM-DD (must be within 7 days; Amman local)")
    p_typ.add_argument("--interval-min", type=int, default=30)
    p_typ.add_argument("--out",          type=Path,
                       default=Path("data/research/gmaps/typical.ndjson"))
    p_typ.set_defaults(func=cmd_typical)

    p_pol = sub.add_parser("poll", help="Live polling daemon (departureTime=now).")
    p_pol.add_argument("--config",     type=Path, default=CONFIG_DEFAULT)
    p_pol.add_argument("--interval-s", type=int, default=300)
    p_pol.add_argument("--out",        type=Path,
                       default=Path("data/events/gmaps_traffic.ndjson"))
    p_pol.add_argument("--once", action="store_true",
                       help="Run a single polling cycle and exit (smoke test)")
    p_pol.set_defaults(func=cmd_poll)

    args = p.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
